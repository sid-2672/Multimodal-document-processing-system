"""
Document ingestion pipeline for RAG system.
Extracts text from various file formats, chunks it, embeds it, and stores in Milvus.
"""

import os
import io
import uuid
import re
import fitz  # PyMuPDF
import pytesseract
from PIL import Image
from docx import Document
from openpyxl import load_workbook
from tqdm import tqdm
from sentence_transformers import SentenceTransformer
from pymilvus import (
    connections,
    FieldSchema,
    CollectionSchema,
    DataType,
    Collection,
    utility
)


def generate_collection_name(file_path: str) -> str:
    """
    Creates a unique collection name based on file extension and name.
    Format: {extension}_{filename}_{random_suffix}
    """
    ext = os.path.splitext(file_path)[1].lower().replace(".", "")
    filename = os.path.basename(file_path).split(".")[0]
    clean_name = re.sub(r"[^a-zA-Z0-9_]+", "_", filename).lower()
    suffix = uuid.uuid4().hex[:6]
    return f"{ext}_{clean_name}_{suffix}"


def init_milvus(collection_name: str, emb_dim: int) -> Collection:
    """
    Sets up a Milvus collection with vector search index.
    Drops existing collection if it already exists.
    """
    connections.connect(alias="default", host="localhost", port="19530")
    
    # Clean slate if collection already exists
    if collection_name in utility.list_collections():
        utility.drop_collection(collection_name)
    
    # Define schema: ID, chunk_id, text content, and embedding vector
    fields = [
        FieldSchema(name="id", dtype=DataType.INT64, is_primary=True, auto_id=True),
        FieldSchema(name="chunk_id", dtype=DataType.INT64),
        FieldSchema(name="text", dtype=DataType.VARCHAR, max_length=65535),
        FieldSchema(name="embedding", dtype=DataType.FLOAT_VECTOR, dim=emb_dim),
    ]
    
    schema = CollectionSchema(fields, description="RAG file chunks")
    collection = Collection(name=collection_name, schema=schema, shards_num=2)
    
    # IVF_FLAT index for fast approximate nearest neighbor search
    index_params = {
        "metric_type": "L2",
        "index_type": "IVF_FLAT",
        "params": {"nlist": 512},
    }
    collection.create_index("embedding", index_params)
    
    return collection


def chunk_text(text: str, max_len: int = 800):
    """
    Splits text into chunks of roughly max_len words.
    Simple word-based chunking, no overlap.
    """
    words = text.split()
    chunk, chunks = [], []
    
    for word in words:
        chunk.append(word)
        if len(chunk) >= max_len:
            chunks.append(" ".join(chunk))
            chunk = []
    
    if chunk:
        chunks.append(" ".join(chunk))
    
    return chunks


def extract_pdf(path: str):
    """
    Extracts text from PDF. Falls back to OCR if a page has no extractable text.
    """
    doc = fitz.open(path)
    extracted_text = []
    
    for page_num in range(doc.page_count):
        page = doc.load_page(page_num)
        text = page.get_text("text").strip()
        
        if text:
            extracted_text.append(text)
            continue
        
        # No text found, try OCR
        pix = page.get_pixmap(dpi=300)
        img = Image.open(io.BytesIO(pix.tobytes("png")))
        ocr_text = pytesseract.image_to_string(img)
        extracted_text.append(ocr_text)
    
    return extracted_text


def extract_docx(path: str):
    """Extracts text from Word documents, paragraph by paragraph."""
    doc = Document(path)
    return [p.text for p in doc.paragraphs if p.text.strip()]


def extract_xlsx(path: str):
    """Extracts text from Excel files, treating each row as a text block."""
    wb = load_workbook(path)
    extracted = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            row_text = " ".join([str(cell) for cell in row if cell is not None])
            if row_text.strip():
                extracted.append(row_text)
    
    return extracted


def extract_image(path: str):
    """Runs OCR on image files to extract text."""
    img = Image.open(path)
    return [pytesseract.image_to_string(img)]


def extract_text_from_file(path: str):
    """Routes to the appropriate extractor based on file extension."""
    ext = os.path.splitext(path)[1].lower()
    
    if ext == ".pdf":
        return extract_pdf(path)
    elif ext in (".doc", ".docx"):
        return extract_docx(path)
    elif ext == ".xlsx":
        return extract_xlsx(path)
    elif ext in (".jpg", ".jpeg", ".png"):
        return extract_image(path)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def ingest_file(file_path: str):
    """
    Main ingestion function. Takes a file, extracts text, chunks it, embeds it,
    and stores everything in Milvus for later retrieval.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    
    collection_name = generate_collection_name(file_path)
    
    # Load embedding model
    model = SentenceTransformer("sentence-transformers/all-mpnet-base-v2")
    emb_dim = len(model.encode(["test"])[0])
    
    collection = init_milvus(collection_name, emb_dim)
    
    # Extract and chunk text
    sections = extract_text_from_file(file_path)
    all_chunks, embeddings, chunk_ids = [], [], []
    chunk_counter = 0
    
    for section in tqdm(sections, desc="Processing chunks"):
        chunks = chunk_text(section)
        if not chunks:
            continue
        
        embs = model.encode(chunks)
        
        for i, chunk in enumerate(chunks):
            all_chunks.append(chunk)
            embeddings.append(embs[i])
            chunk_ids.append(chunk_counter)
            chunk_counter += 1
    
    # Insert into Milvus
    collection.insert([chunk_ids, all_chunks, embeddings])
    collection.flush()
    collection.load()
    
    return {
        "collection": collection_name,
        "chunks": len(all_chunks),
    }


def query_milvus(collection_name: str, question: str):
    """
    Searches Milvus for the top 5 most relevant text chunks for a given question.
    """
    model = SentenceTransformer("sentence-transformers/all-mpnet-base-v2")
    query_vec = model.encode([question])[0]
    
    col = Collection(collection_name)
    results = col.search(
        data=[query_vec],
        anns_field="embedding",
        param={"metric_type": "L2", "params": {"nprobe": 16}},
        limit=5,
        output_fields=["text"]
    )
    
    return [hit.entity.get("text") for hit in results[0]]

